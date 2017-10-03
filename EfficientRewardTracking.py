'''
RewardTracker that runs once at UTC1200 and gets the Curation history of Curators over 7 days compared with Curie's curation rewards

Made by @locikll

Packages: Piston-lib, Piston-cli, steem-python

'''

import sys
import datetime
import os
import subprocess
import math
import re
import csv
import time

from openpyxl import *
import pickle

from time import gmtime, strftime
import os.path
from pathlib import Path
import timeit

import multiprocessing.dummy as mp 
from multiprocessing import process
from collections import OrderedDict
import websocket


import piston
import steem as pysteem
from piston.steem import Steem
from piston.utils import parse_time
from piston.amount import Amount
from random import randint

from datetime import datetime as dti

#EDITABLE VARIABLES {{

#Steemit Node & Account stuff
webs = websocket.WebSocket()

Node = ["wss://steemd.steemit.com","wss://this.piston.rocks","wss://gtg.steem.house:8090","wss://seed.bitcoiner.me"]

steemPostingKey = ''  #Private posting key
steemAccountName = 'locikll' 

#Number of days to go back from current
tday = 7

#Reward Limit,  Increase this for Increased number of days (This is to save computational resources, this can be made very large without much time increment (Due to loop breaking after time limit), but will cost more memory)
REWARDLIM = 10000

curationhistaccount = 'curie' #Account for reward history

#List of Followed Curators
followedcurators = ['secret','somecuratorname2']

# EDITABLE VARIABLES STOP  }}

curatordict = OrderedDict((el,[]) for el in followedcurators)

#trailing_7d_t = time.time() - datetime.timedelta(days=7).total_seconds()

starttime = datetime.datetime.now()
daystotrail = datetime.timedelta(tday,0,0)

#mindaystrail = datetime.timedelta(2,0,0)

#for curator in range(0,len(followedcurators)):
    
#curhist = piston.account.Account(curationhistaccount).history2(filter_by="curation_reward", take=10000)
curhist = pysteem.account.Account(curationhistaccount).get_account_history(filter_by="curation_reward",limit=REWARDLIM,index=-1,order=-1)
#Initial File Manipulation and storage
wb = Workbook()
ws = wb.active
filepath = ( 'reports/'+str(dti.now().day)+'.'+str(dti.now().month)+'.'+str(dti.now().year) )

def setupfiledir():

        if not os.path.exists(filepath):
                os.makedirs(filepath)

        for Files in range(0,len(followedcurators)):

                filenamepath = filepath+'/'+followedcurators[Files]+'.xlsx'
                curfiles = Path(filenamepath)

                if not curfiles.is_file():

                        ws['A1'] = 'Post ID'
                        ws['B1'] = 'Date'
                        ws['C1'] = 'VP (%)'
                        ws['D1'] = 'Curation Reward (STEEM)'
                        ws['E1'] = 'Running Curation Reward Total (STEEM)'
                        ws['F1'] = 'Total Reward (STEEM)'

                        wb.save(filename=filenamepath)

        if not Path(filepath+'/'+'TOTALS.xlsx').is_file():
                ws['A1'] = 'Curator Account'
                ws['B1'] = 'Total VP Used (%)'
                ws['C1'] = 'Total Curation Rewards Received'
                wb.save(filename=filepath+'/'+'TOTALS.xlsx')

#Check that Websocket is connected, if not try another one
def selectnode(NodeNumber):
        #Try the websocket, if it fails then print fail and try another websocket
        try:
                webs.connect(Node[NodeNumber]) 
                print('connected to node:' + Node[NodeNumber])
        except Exception:
                print(Exception)
                NodeNumber=NodeNumber+1
                selectnode(NodeNumber)

#Make sure Nodenumber doesn't exceed number of nodes
        if NodeNumber > (len(Node)-1):
                NodeNumber=0 

        steem = Steem(wif=steemPostingKey,node=Node[NodeNumber])
        
        return NodeNumber

def curator_rewards(steem,NodeNumber):    
        
        steemperMvest = pysteem.account.Account(account_name=curationhistaccount).converter.steem_per_mvests()
        n = 0
                
        for reward in curhist:
                
                #print(reward)
                n = n+1
                print(n)
        
                timestamp = parse_time(reward['timestamp'])                
                timediff = starttime - timestamp
       
                #If post is within the past 7 days
                if daystotrail >= timediff:
                                               
                        author = reward['comment_author']
                        permlink = reward['comment_permlink']
                        identifier = '@'+author+'/'+permlink
                
                        postmeta = GET_post(identifier)
                        activevotes = postmeta[2]                              
                                                
                        #Sort votes by datetime
                        if activevotes != 'ERROR':
                                activevotes.sort(key=lambda item:item['time'], reverse=False)              
                                voters = [voter['voter'] for voter in activevotes]
                                curatorsvoted = list(set(voters).intersection(followedcurators))
                        else:
                                curatorsvoted = []
                
                        if curatorsvoted != []:
                                
                                print('Doing Steem account.account')
                                precurievoter = curatorsvoted[0]
                                votetime = (list(filter(lambda voter: voter['voter']==precurievoter,activevotes))[0]['time']).replace('T',' ')
                                
                                
                                
                                curreward = (Amount(reward['reward']).amount / 1e6) * steemperMvest
                                
                                votingpower = (list(filter(lambda voter: voter['voter']==curationhistaccount,activevotes))[0]['percent'])/100
                        
                                docdata = [identifier,votetime,votingpower,curreward]
                                                                
                                curatordict[precurievoter].append(docdata)
                                
                                print(docdata)
                           
                        else:
                                continue
 
                else:
                        break
                
        
        #After done, save to file(s)
        print('saving files...')
        for cur in range(0,len(followedcurators)):
                
                #Calculate Total rewards for each user and save                
                wbtot = load_workbook(filepath+'/'+'TOTALS.xlsx')
                wstot = wbtot.active
                
                VPtotal = sum( [ el[2] for el in curatordict[followedcurators[cur]][0:len(curatordict[followedcurators[cur]])] ] )
                RewardTotal = sum( [ es[3] for es in curatordict[followedcurators[cur]][0:len(curatordict[followedcurators[cur]])] ] )
                
                totaldoc = [followedcurators[cur],VPtotal,RewardTotal]
                
                wstot.append(totaldoc)
                wbtot.save(filepath+'/'+'TOTALS.xlsx')
                
                #Calculate and save Individual rewards for posts and save
                wb = load_workbook(filepath+'/'+followedcurators[cur]+'.xlsx')
                ws = wb.active                  
        
                for savepost in range(0,len(curatordict[followedcurators[cur]])):
             
                        ws.append( (curatordict[followedcurators[cur]])[savepost] )

                        lastrow = ws.max_row
        
                        ws[("E"+str(lastrow))] = "=SUM(D1:"+"D"+str(lastrow)+")"           
                        ws["F2"] = "=SUM(D:D)"
                        
        
                wb.save(filepath+'/'+followedcurators[cur]+'.xlsx')    
                                
        print('saved')
                
def GET_post(identifier):

        try:
                postid = steem.get_post(identifier)
                posttitle = postid.title        
                postvotes = postid.active_votes
                postauthor = postid.author                
        except Exception:
                print('Exception occured with Identifier: '+identifier)
                postid = 'ERROR'
                posttitle = 'ERROR'
                postvotes = 'ERROR'
                postauthor = 'ERROR'
                pass
                
        return [postid,posttitle,postvotes,postauthor]

if __name__ == "__main__":
        NodeNumber = 2
        
        #Run initial function for setting up directories/checking files            
        setupfiledir()          
        
        #Run Select Node with Initial Node Number        
        NodeNumber = selectnode(NodeNumber)
        
        steem = Steem(wif=steemPostingKey,node=Node[NodeNumber])
        
        curator_rewards(steem,NodeNumber)